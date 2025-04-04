import pandas as pd
import re
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from logsApp.models import RegistredCars, EmployesInfo, InUseCars, LogsC, AccidentsRecord, FinesAccidentsImage, LicenseFile, FinesRecord
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta
from django.db.models import Q
from django.db import IntegrityError
import pytz
from django.core.paginator import Paginator
import shutil
import os
from .cars import carsList
from .newempData import newemp


def seedemp(request):
    for i in newemp:
        emp, created = EmployesInfo.objects.update_or_create(
            ceoNumber=i["empId"],
            defaults={
                'ceoName': i["empName"],
                'phoneNumber': i["tel"],
                'jobTtile': i["jobTitle"],
                'department': i["department"],
                'unit': i["unit"],
                'nationality': i["nationality"]
            }
        )

    for q in carsList:
        car, created = RegistredCars.objects.update_or_create(
            carNumber=q["vnumber"],
            defaults={
                'vType': q["vType"],
                'carYear': q['Myear'],
                'cownerEmpNumber': q['empid'],
                'cownerName': q['empName'],          
            }
        )

    return HttpResponse("done")


def remove_non_numeric(s):
    return re.sub(r'\D', '', s)

def removechars(c):
    characters_to_remove = "!@#$%^&*()_+|?}{:><`1234567890"

    for char in characters_to_remove:
        c = c.replace(char, "")
    
    return c

def index(request):
    try:
        # Get all in-use cars for the dashboard
        all_in_use_cars = InUseCars.objects.select_related('car', 'employee').all().order_by('-id')
        return render(request, "logsApp/registerCar.html", {"l": all_in_use_cars})
    except Exception as e:
        print(f"Error in index view: {str(e)}")
        # Return a simple error page if something goes wrong
        return render(request, "logsApp/error.html", {"error_message": "An error occurred. Please try again."})

def registerCar(request):
    all_in_use_cars = InUseCars.objects.all().order_by('-id')
    
    if request.method == "POST":
        ceo_number = request.POST.get("ceoNumber").strip()
        car_number = request.POST.get("carNumber").strip()

        try:
            its_in_use = LogsC.objects.get(Logs_employee_ins__ceoNumber=ceo_number, carIsInUse=True)
            return render(request, "logsApp/registerCar.html", {
                "itsinuse": its_in_use,
                "l": all_in_use_cars,
                "ceoNumber": "",
                "carNumber": car_number,
                "form_open": True
            })
        except LogsC.DoesNotExist:
            pass

        try:
            car_is_in_use = LogsC.objects.get(Logs_car_ins__carNumber=car_number, carIsInUse=True)
            return render(request, "logsApp/registerCar.html", {
                "carIsInUse": car_is_in_use,
                "l": all_in_use_cars,
                "ceoNumber": ceo_number,
                "carNumber": "",
                "form_open": True
            })
        except LogsC.DoesNotExist:
            pass

        try:
            emp_exists = EmployesInfo.objects.get(ceoNumber=ceo_number, EmpHaveCar=False)
        except EmployesInfo.DoesNotExist:
            emp_does_not_exist = "الرقم الاداري غير صحيح"
            return render(request, "logsApp/registerCar.html", {
                "empDoseNotEXISTS": emp_does_not_exist,
                "l": all_in_use_cars,
                "ceoNumber": "",
                "carNumber": car_number,
                "form_open": True
            })

        try:
            car_exists = RegistredCars.objects.get(carNumber=car_number, carIsInparking=True)
        except RegistredCars.DoesNotExist:
            car_dne = "رقم المركبة غير صحيح"
            return render(request, "logsApp/registerCar.html", {
                "carDNE": car_dne,
                "l": all_in_use_cars,
                "ceoNumber": ceo_number,
                "carNumber": "",
                "form_open": True
            })

        InUseCars.objects.create(car=car_exists, employee=emp_exists)
        LogsC.objects.create(Logs_employee_ins=emp_exists, Logs_car_ins=car_exists)

        emp_exists.EmpHaveCar = True
        emp_exists.save()
        car_exists.carIsInparking = False
        car_exists.save()

        success_message = "تم التسجيل بنجاح"
        return render(request, "logsApp/registerCar.html", {
            "sucssuMessge": success_message,
            "l": all_in_use_cars
        })

    return render(request, "logsApp/registerCar.html", {"l": all_in_use_cars})

def returnCar(request):
    if request.method == "POST":
        ceo_number = remove_non_numeric(request.POST.get("ceonumber")).strip()
        emp_note = request.POST.get("empnote")
        all_in_use_cars = InUseCars.objects.select_related('car', 'employee').all().order_by('-id')
        dubai_tz = pytz.timezone('Asia/Dubai')

        try:
            emp_instance = EmployesInfo.objects.get(ceoNumber=ceo_number)
        except EmployesInfo.DoesNotExist:
            ret_err_msg = "الرقم الاداري غير صحيح"
            return render(request, "logsApp/registerCar.html",{
                "retErrm": ret_err_msg,
                "l": all_in_use_cars,
                "ceonumber": "",
                "empnote": "emp_note",
                "form_open": True
            })

        try:
            in_use_car_instance = InUseCars.objects.select_related('car', 'employee').get(employee=emp_instance)
        except InUseCars.DoesNotExist:
            ret_car_err = "لاتوجد مركبه مرتبطه بل رقم الاداري"
            return render(request, "logsApp/registerCar.html", {
                "retCarErr": ret_car_err,
                "l": all_in_use_cars,
                "ceonumber": "",
                "empnote": "emp_note",
                "form_open": True})

        ret_success_msg = "تم اعاده المركبه بنجاح"
        registered_car_instance = RegistredCars.objects.get(carNumber=in_use_car_instance.car.carNumber)
        log_instance = LogsC.objects.select_related('Logs_employee_ins', 'Logs_car_ins').get(Logs_employee_ins=emp_instance, carIsInUse=True)

        current_time = timezone.now().astimezone(dubai_tz)
        log_instance.ended_at = current_time
        log_instance.return_date = current_time.date()
        log_instance.return_time = current_time.time()
        log_instance.carIsInUse = False
        log_instance.carNote = emp_note

        registered_car_instance.carIsInparking = True
        emp_instance.EmpHaveCar = False

        emp_instance.save()
        log_instance.save()
        registered_car_instance.save()
        in_use_car_instance.delete()

        return render(request, "logsApp/registerCar.html", {"retSucssM": ret_success_msg,"l": all_in_use_cars})

    return render(request, "logsApp/registerCar.html", {"l": InUseCars.objects.select_related('car', 'employee').all().order_by('-id')})

def logsfunc(request):
    years = range(2020, 2040)
    if request.method == "POST" or request.method == "GET":
        car_number = request.GET.get('carNumper')
        ceo_number = request.GET.get('ceoN')
        date_filter = request.GET.get('date')
        month_filter = request.GET.get('month')
        year_filter = request.GET.get('year')
        year_only_filter = request.GET.get('yearOnly')
        date_type = request.GET.get('dateType')
        show_all = request.GET.get('showAll')

        filters = Q()

        if date_type == "day" and date_filter:
            filters &= Q(taken_date=date_filter)
        elif date_type == "month" and month_filter and year_filter:
            month_start = datetime(year=int(year_filter), month=int(month_filter), day=1)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(seconds=1)
            filters &= Q(taken_date__range=(month_start, month_end))
        elif date_type == "year" and year_only_filter:
            year_start = datetime(year=int(year_only_filter), month=1, day=1)
            year_end = datetime(year=int(year_only_filter), month=12, day=31, hour=23, minute=59, second=59)
            filters &= Q(taken_date__range=(year_start, year_end))

        if ceo_number:
            filters &= Q(Logs_employee_ins__ceoNumber=ceo_number.strip())

        if car_number:
            filters &= Q(Logs_car_ins__carNumber=car_number.strip())

        if show_all:
            logs = LogsC.objects.all().order_by('-id')
        else:
            logs = LogsC.objects.filter(filters).order_by('-id')


        paginator = Paginator(logs, 50)  
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        return render(request, "logsApp/logs.html", {'page_obj': page_obj, 'years': years})

    current_date = datetime.now().date()
    today_logs = LogsC.objects.filter(Q(taken_date=current_date) | Q(taken_date__isnull=True)).order_by('-id')
    
    paginator = Paginator(today_logs, 50) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, "logsApp/logs.html", {'page_obj': page_obj, 'years': years})

def is_pdf(file_field):
    return file_field.url.endswith(".pdf")

def AccidentsRecords(request):
    fines = AccidentsRecord.objects.all()
    if request.method == "POST":
        try:
            car_number = request.POST.get('carNumber')
            emp_number = request.POST.get('empNumber')
            text = request.POST.get('text')
            report_pdf_file = request.FILES.get('reportPdfFile')
            car_paperwork_file = request.FILES.get('carPaperworkFile')
            license_files = request.FILES.getlist('licenseFiles')
            images = request.FILES.getlist('images')

            try:
                car_instance = RegistredCars.objects.get(carNumber=car_number)
            except RegistredCars.DoesNotExist:
                car_instance = None
                
            emp_instance = None
            if emp_number:
                try:
                    emp_instance = EmployesInfo.objects.get(ceoNumber=emp_number)
                except EmployesInfo.DoesNotExist:
                    emp_err_message = "الرقم الاداري غير صحيح"
                    return render(request, "logsApp/accidentsPage.html", {
                        'form_open': True,
                        'fines': fines,
                        'emp_err_message': emp_err_message,
                        'images': images,
                        'carNumber': car_number,
                        'empNumber': emp_number,
                        'text': text,
                        'reportPdfFile': report_pdf_file,
                        'carPaperworkFile': car_paperwork_file,
                        'licenseFiles': license_files
                    })

            accidents_record = AccidentsRecord.objects.create(
                car=car_instance,
                text=text)

            if report_pdf_file:
                accidents_record.report_pdf_file = report_pdf_file
            if car_paperwork_file:
                accidents_record.car_paperwork_file = car_paperwork_file

            if emp_instance:
                accidents_record.employees.add(emp_instance)

            for image in images:
                FinesAccidentsImage.objects.create(accidents_record=accidents_record, image=image)

            for license_file in license_files:
                LicenseFile.objects.create(accidents_record=accidents_record, file=license_file)

            accidents_record.save()
        except Exception as e:
            print(f"Error in AccidentsRecords view: {str(e)}")
            # You might want to add error handling here

    return render(request, "logsApp/accidentsPage.html", {'fines': fines, 'form_open': False})

def export_to_excel(request):
    dubai_tz = pytz.timezone('Asia/Dubai')
    
    data = LogsC.objects.select_related('Logs_employee_ins', 'Logs_car_ins').all().order_by('-id')

    export_data = []
    for log in data:
        created_at_dubai = log.created_at.astimezone(dubai_tz)
        taken_time_dubai = (log.taken_time.replace(tzinfo=dubai_tz) if log.taken_time else None)
        ended_at_dubai = log.ended_at.astimezone(dubai_tz) if log.ended_at else None
        return_time_dubai = (log.return_time.replace(tzinfo=dubai_tz) if return_time_dubai else None)

        export_data.append({
            'ID': log.id,
            'رقم السياره': log.Logs_car_ins.carNumber,
            'الاسم': log.Logs_employee_ins.ceoName,
            'الرقم الاداري': log.Logs_employee_ins.ceoNumber,
            'تاريخ ووقت الاستلام': created_at_dubai.strftime('%Y-%m-%d %I:%M %p'),
            'تاريخ الاستلام': log.taken_date,
            'وقت الاستلام': taken_time_dubai.strftime('%I:%M %p') if taken_time_dubai else None,
            'تاريخ و وقت التسليم': ended_at_dubai.strftime('%Y-%m-%d %I:%M %p') if ended_at_dubai else None,
            'تاريخ التسليم': log.return_date,
            'وقت التسليم': return_time_dubai.strftime('%I:%M %p') if return_time_dubai else None,
            'ملاحظه على المركبه': log.carNote,
            'قسم الموظف': log.Logs_employee_ins.section
        })
    
    df = pd.DataFrame(export_data)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="logs_data.xlsx"'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Logs Data')

        workbook = writer.book
        worksheet = writer.sheets['Logs Data']
       
        worksheet.sheet_view.rightToLeft = True
       
        header_font = Font(size=16, bold=True, color='000000')
        header_fill = PatternFill(start_color='B7E1A1', end_color='B7E1A1', fill_type='solid')
        cell_font = Font(size=16)
        center_alignment = Alignment(horizontal='center')

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
       
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.font = cell_font
                cell.alignment = center_alignment
       
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 4)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    return response

def addNewEmp(request):
    if request.method == "POST":
        empNameq = request.POST.get('empName')
        empNumber = request.POST.get('empNumber')
        sa = removechars(empNameq)
        print(sa)
        try:
            EmployesInfo.objects.create(ceoName=sa,ceoNumber=empNumber)
            dd = " working"
        except IntegrityError:
            dd = "الرقم الاداري موجود من قبل"
        return render(request, "logsApp/addNewEmp.html", {'dd': dd})
    return render(request,"logsApp/addNewEmp.html")

def fineC(request):
    finon = None
    allFines = FinesRecord.objects.all()
    if request.method == "POST":
        fine_date = request.POST.get('finedate')
        fine_time = request.POST.get('finetime')
        fine_car_number = request.POST.get('finecar')
        dubai_tz = pytz.timezone('Asia/Dubai')
        combined_fine_datetime = dubai_tz.localize(timezone.datetime.strptime(f"{fine_date} {fine_time}", '%Y-%m-%d %H:%M'))
        
        try:
            car_ins = RegistredCars.objects.get(carNumber=fine_car_number)

            finon = LogsC.objects.get(Logs_car_ins=car_ins,
                                      created_at__lte = combined_fine_datetime,
                                      ended_at__gte = combined_fine_datetime
                                      )
            # Save the fine details in the FinesRecord model
            FinesRecord.objects.create(
                car=car_ins,
                employe=finon.Logs_employee_ins,
                created_at=combined_fine_datetime
            )               
            return redirect('logsApp:finespage')
        except RegistredCars.DoesNotExist:
            print(f"Car with number {fine_car_number} does not exist.")
        except Exception as e:
            print(f"An error occurred: {e}")

        return render(request, "logsApp/finespage.html", {'finon': finon})
    return render(request, "logsApp/finespage.html", {'allFines':allFines,'finon': finon})
    
def carddetails(request, fine_id):
    fine = get_object_or_404(AccidentsRecord, id=fine_id)
    if request.method == "POST":
        try:
            report_pdf_file = request.FILES.get('reportPdfFile')
            car_paperwork_file = request.FILES.get('carPaperworkFile')
            license_files = request.FILES.getlist('licenseFiles')
            images = request.FILES.getlist('images')
            emp_number = request.POST.get('empNumber')

            print(f"Processing files for fine_id: {fine_id}")
            print(f"Report PDF: {report_pdf_file}")
            print(f"Car Paperwork: {car_paperwork_file}")
            print(f"License Files: {len(license_files)}")
            print(f"Images: {len(images)}")

            if report_pdf_file:
                try:
                    fine.report_pdf_file = report_pdf_file
                    print("Successfully saved report PDF")
                except Exception as e:
                    print(f"Error saving report PDF: {str(e)}")

            if car_paperwork_file:
                try:
                    fine.car_paperwork_file = car_paperwork_file
                    print("Successfully saved car paperwork")
                except Exception as e:
                    print(f"Error saving car paperwork: {str(e)}")

            for image in images:
                try:
                    FinesAccidentsImage.objects.create(accidents_record=fine, image=image)
                    print(f"Successfully saved image: {image.name}")
                except Exception as e:
                    print(f"Error saving image {image.name}: {str(e)}")

            for license_file in license_files:
                try:
                    LicenseFile.objects.create(accidents_record=fine, file=license_file)
                    print(f"Successfully saved license file: {license_file.name}")
                except Exception as e:
                    print(f"Error saving license file {license_file.name}: {str(e)}")

            if emp_number:
                try:
                    emp_instance = EmployesInfo.objects.get(ceoNumber=emp_number)
                    fine.employees.add(emp_instance)
                    print(f"Successfully added employee: {emp_number}")
                except EmployesInfo.DoesNotExist:
                    print(f"Employee not found: {emp_number}")
                except Exception as e:
                    print(f"Error adding employee: {str(e)}")

            try:
                fine.save()
                print("Successfully saved fine record")
            except Exception as e:
                print(f"Error saving fine record: {str(e)}")

        except Exception as e:
            print(f"Error in carddetails view: {str(e)}")
            # You might want to add error handling here

    license_files = []
    license_images = []
    for license_file in fine.license_files.all():
        if is_pdf(license_file.file):
            license_files.append(license_file)
        else:
            license_images.append(license_file)
    return render(request, 'logsApp/carddetails.html', {
        'fine': fine,
        'license_files': license_files,
        'license_images': license_images
    })

def markasfixed(request, fine_id):
    fine = get_object_or_404(AccidentsRecord, id=fine_id)
    fine.fixin_date = timezone.now()

    fine.save()
    return redirect('logsApp:carddetails', fine_id=fine_id)

def fineDetails(request, fine_id):
    fine = get_object_or_404(FinesRecord, id=fine_id)
    fine_files = []
    fine_images = []
    if fine.paid_fine_image:
        if is_pdf(fine.paid_fine_image):
            fine_files.append(fine.paid_fine_image)
        else:
            fine_images.append(fine.paid_fine_image)
    if request.method == "POST":
        paid_fine_image = request.FILES.get('paidFineImage')
        if paid_fine_image:
            fine.paid_fine_image = paid_fine_image
            fine.paidDate = timezone.now().date()
            fine.save()
            fine = get_object_or_404(FinesRecord, id=fine_id)
            fine_files = []
            fine_images = []
            if fine.paid_fine_image:
                if is_pdf(fine.paid_fine_image):
                    fine_files.append(fine.paid_fine_image)
                else:
                    fine_images.append(fine.paid_fine_image)
    return render(request, 'logsApp/fineDetails.html', {'fine': fine, 'fine_files': fine_files, 'fine_images': fine_images})

def deleteFineImage(request, fine_id):
    fine = get_object_or_404(FinesRecord, id=fine_id)
    if fine.paid_fine_image:
        # Clear the fields in the database
        # The S3 storage backend will handle file deletion automatically
        fine.paid_fine_image = None
        fine.paidDate = None
        fine.save()
    return redirect('logsApp:fineDetails', fine_id=fine_id)